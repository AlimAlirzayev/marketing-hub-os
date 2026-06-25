"""Import sold-policy rows from a CRM export (CSV) and send them to Meta as
Purchase events on the offline dataset — no developer, no CRM API needed.

The marketing/sales team just exports the month's issued policies and runs this.
Each row becomes a Purchase with the premium as `value`; the policy number is
the dedup key, so re-running the same export never double-counts.

Safe by default — it only previews. Add --send to actually send.

    # 1) Preview what would be sent (hashes PII, sends nothing):
    .venv\\Scripts\\python.exe import_sales.py sales.csv

    # 2) Send into Test Events first (set META_TEST_EVENT_CODE or pass --test-code):
    .venv\\Scripts\\python.exe import_sales.py sales.csv --send --test-code TEST12345

    # 3) Go live:
    .venv\\Scripts\\python.exe import_sales.py sales.csv --send

Expected columns (Azerbaijani/English headers auto-detected; override with
--map "value=Premium,policy_no=Polis No"):
    policy_no*  premium*  currency  product  email  phone
    first_name  last_name  date  external_id
(* required)
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
import time
from datetime import date, datetime

import capi
import config


class UnsupportedFile(Exception):
    """Raised with a user-friendly message when a file can't be read."""

# Canonical field -> accepted header forms (matched case/space/punct-insensitive).
_ALIASES: dict[str, list[str]] = {
    "policy_no":  ["policyno", "policy", "polis", "polisno", "polisnomresi",
                   "muqavile", "muqaveleno", "contract", "orderid", "order"],
    "premium":    ["premium", "value", "amount", "mebleg", "deyer", "sum",
                   "qiymet", "mblg", "price"],
    "currency":   ["currency", "valyuta", "ccy"],
    "product":    ["product", "mehsul", "contentname", "nov", "mehsuladi", "tarif"],
    "email":      ["email", "mail", "epoct", "emailaddress"],
    "phone":      ["phone", "telefon", "mobil", "mobile", "nomre", "number", "gsm"],
    "first_name": ["firstname", "ad", "name", "musteriadi"],
    "last_name":  ["lastname", "soyad", "surname"],
    "date":       ["date", "saledate", "tarix", "satistarixi", "satis", "eventtime"],
    "external_id": ["externalid", "customerid", "musteriid", "clientid", "fin"],
}

_DATE_FORMATS = ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d",
                 "%Y-%m-%dT%H:%M:%S", "%d.%m.%Y %H:%M", "%m/%d/%Y")


def _norm(h: str) -> str:
    """Lowercase a header and strip everything but letters/digits (AZ-safe)."""
    h = (h or "").strip().lower()
    h = (h.replace("ə", "e").replace("ı", "i").replace("ğ", "g")
           .replace("ş", "s").replace("ç", "c").replace("ö", "o").replace("ü", "u"))
    return re.sub(r"[^a-z0-9]", "", h)


def detect_columns(headers: list[str], overrides: dict[str, str]) -> dict[str, str]:
    """Map canonical field -> actual header. Manual overrides win."""
    norm_to_real = {_norm(h): h for h in headers}
    mapping: dict[str, str] = {}
    for field, aliases in _ALIASES.items():
        if field in overrides:
            mapping[field] = overrides[field]
            continue
        for alias in aliases:
            # Exact match, or a long alias contained in a header (e.g. "məbləğ" in "Məbləğ (AZN)").
            hit = norm_to_real.get(alias)
            if not hit and len(alias) >= 5:
                hit = next((real for n, real in norm_to_real.items() if alias in n), None)
            if hit:
                mapping[field] = hit
                break
    return mapping


def parse_date(raw: str) -> int | None:
    raw = (raw or "").strip()
    if not raw:
        return None
    if raw.isdigit() and len(raw) >= 9:        # already a unix timestamp
        return int(raw)
    for fmt in _DATE_FORMATS:
        try:
            return int(datetime.strptime(raw, fmt).timestamp())
        except ValueError:
            continue
    return None


def parse_amount(raw: str) -> float | None:
    s = re.sub(r"[^\d,.\-]", "", (raw or "").strip())
    if not s:
        return None
    # Treat the last separator as the decimal point; drop the rest (thousands).
    s = s.replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(",", "") if s.rfind(".") > s.rfind(",") else s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def _cell_to_str(c) -> str:
    """Normalise any spreadsheet cell to a clean string."""
    if c is None:
        return ""
    if isinstance(c, datetime):
        return c.strftime("%Y-%m-%d")
    if isinstance(c, date):
        return c.isoformat()
    if isinstance(c, float) and c.is_integer():   # 994501234567.0 -> "994501234567"
        return str(int(c))
    return str(c).strip()


def _grid_xlsx(path: str) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise UnsupportedFile(
            "Excel oxumaq üçün 'openpyxl' lazımdır (bir dəfəlik). Quraşdır:\n"
            "       .venv\\Scripts\\python.exe -m pip install openpyxl\n"
            "    — və ya Excel-də 'Save As' → 'CSV UTF-8' seçib CSV ver.")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    grid = [[_cell_to_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    return grid


def _grid_csv(path: str) -> list[list[str]]:
    # AZ Excel often saves CSV as cp1254/latin-1 and with ';' separators.
    for enc in ("utf-8-sig", "cp1254", "latin-1"):
        try:
            with open(path, encoding=enc, newline="") as f:
                sample = f.read(8192)
                f.seek(0)
                delim = ";" if sample.count(";") > sample.count(",") else ","
                return [[(c or "").strip() for c in row]
                        for row in csv.reader(f, delimiter=delim)]
        except UnicodeDecodeError:
            continue
    raise UnsupportedFile(
        "CSV oxunmadı (kodlaşdırma problemi). Excel-də 'Save As' → "
        "'CSV UTF-8' seçib yenidən yadda saxla.")


def _find_header(grid: list[list[str]], overrides: dict[str, str], max_scan: int = 15) -> int:
    """Pick the most likely header row — exports often prefix a title/filter banner."""
    best_i, best_score = 0, -1
    for i in range(min(max_scan, len(grid))):
        cells = [c for c in grid[i] if c]
        if len(cells) < 3:
            continue
        m = detect_columns(cells, overrides)
        score = len(m) + (10 if "policy_no" in m else 0)
        if score > best_score:
            best_i, best_score = i, score
    return best_i


def read_table(path: str, overrides: dict[str, str]) -> tuple[list[str], list[dict], int]:
    """Read CSV/Excel into (headers, row-dicts, header_row_number). Auto-skips
    any banner rows above the real header."""
    if not os.path.exists(path):
        raise UnsupportedFile(f"Fayl tapılmadı: {path}")
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        grid = _grid_xlsx(path)
    elif ext == ".xls":
        raise UnsupportedFile(
            "Köhnə '.xls' dəstəklənmir. Excel-də 'Save As' → '.xlsx' "
            "və ya 'CSV UTF-8' seçib yenidən yadda saxla.")
    else:
        grid = _grid_csv(path)

    if not grid:
        return [], [], 0
    hi = _find_header(grid, overrides)
    headers = grid[hi]
    rows = []
    for raw in grid[hi + 1:]:
        if not any(c for c in raw):
            continue
        rows.append({h: v for h, v in zip(headers, raw) if h})
    return [h for h in headers if h], rows, hi + 1


def load_rows(rows: list[dict], mapping: dict[str, str],
              max_age_days: int = 62) -> tuple[list[dict], list[str]]:
    """Turn raw table rows into built Purchase events. Returns (events, warnings).

    Meta's offline dataset rejects events older than ~62 days, and one bad event
    fails its whole 1000-event batch, so we skip too-old rows up front.
    """
    events, warnings = [], []
    default_ccy = "AZN"
    no_ident = too_old = 0
    cutoff = int(time.time()) - max_age_days * 86400 if max_age_days else 0
    for i, row in enumerate(rows, start=1):
        def cell(field: str) -> str:
            col = mapping.get(field)
            return (row.get(col) or "").strip() if col else ""

        policy_no = cell("policy_no")
        if not policy_no:
            warnings.append(f"məlumat sətri {i}: polis № yoxdur — ötürüldü")
            continue

        sale_time = parse_date(cell("date"))
        if cutoff and sale_time and sale_time < cutoff:
            too_old += 1
            continue
        premium = parse_amount(cell("premium"))   # None => Purchase without value

        ident = {k: cell(k) for k in ("email", "phone", "first_name",
                                      "last_name", "external_id") if cell(k)}
        ev = capi.build_policy_sale(
            premium=premium, policy_no=policy_no,
            product=cell("product") or None,
            currency=cell("currency") or default_ccy,
            sale_time=sale_time,
            **ident)
        if not ev["user_data"]:   # no usable identifier (Meta rejects such events)
            no_ident += 1
            continue
        events.append(ev)
    if too_old:
        warnings.append(f"{too_old} polis {max_age_days} gündən köhnədir — ötürüldü "
                        "(Meta offline dataset köhnə hadisələri qəbul etmir)")
    if no_ident:
        warnings.append(f"{no_ident} sətirdə yararlı müştəri identifikatoru (email/telefon) "
                        "yoxdur — ötürüldü (Meta belə hadisələri qəbul etmir)")
    return events, warnings


def _chunks(seq: list, n: int):
    for i in range(0, len(seq), n):
        yield seq[i:i + n]


def main() -> int:
    ap = argparse.ArgumentParser(description="CRM policy sales -> Meta Purchase events")
    ap.add_argument("file", help="path to the CRM export (CSV or Excel .xlsx)")
    ap.add_argument("--send", action="store_true", help="actually send (default: preview only)")
    ap.add_argument("--test-code", default=None, help="route to Events Manager → Test Events")
    ap.add_argument("--map", default="", help='manual columns, e.g. "value=Premium,policy_no=Polis No"')
    ap.add_argument("--dataset", default=None, help="override dataset id (default: offline)")
    ap.add_argument("--max-age-days", type=int, default=62,
                    help="skip sales older than this (Meta offline limit ~62; 0 = no limit)")
    args = ap.parse_args()

    overrides = {}
    for pair in filter(None, (p.strip() for p in args.map.split(","))):
        if "=" in pair:
            k, v = (s.strip() for s in pair.split("=", 1))
            overrides[k] = v

    print("=" * 64)
    print("  Meta CAPI · CRM polis satışları → Purchase")
    print("=" * 64)

    try:
        headers, raw_rows, header_line = read_table(args.file, overrides)
    except UnsupportedFile as exc:
        print(f"\n  ✗ {exc}\n")
        return 1

    mapping = detect_columns(headers, overrides)
    print(f"  Fayl     : {args.file}")
    print(f"  Başlıq sətri : {header_line}  ·  məlumat sətri: {len(raw_rows)}")
    print(f"  Sütunlar : {headers}")
    print("  Uyğunlaşma (lazımi sahə → faylın sütunu):")
    for field in ("policy_no", "premium", "currency", "product", "email",
                  "phone", "first_name", "last_name", "date", "external_id"):
        print(f"     {field:<12} → {mapping.get(field, '—')}")
    if "policy_no" not in mapping:
        print("\n  ✗ Vacib sütun tapılmadı: 'policy_no' (polis №). --map ilə göstər, məs:")
        print('     --map "policy_no=Polisi No"\n')
        return 2
    if "premium" not in mapping:
        print("\n  ⚠ 'premium' (məbləğ) sütunu yoxdur → satışlar value=0 ilə göndəriləcək.")
        print("    Konversiya kimi sayılır (alqoritm optimallaşır), amma ROAS/gəlir 0 görünər.")
        print("    Tam fayda üçün ixraca premium sütunu əlavə et (məs. 'Brutto Mükafat').")

    events, warnings = load_rows(raw_rows, mapping, args.max_age_days)
    dataset = args.dataset or (config.OFFLINE_DATASET_ID or config.active_dataset())
    valued = [e for e in events if e["custom_data"].get("value", 0) > 0]
    total = round(sum(e["custom_data"]["value"] for e in valued), 2)
    ccy = valued[0]["custom_data"]["currency"] if valued else "AZN"

    ds_kind = "offline" if dataset == config.OFFLINE_DATASET_ID else "pixel"
    print(f"\n  Hazırlanan hadisə : {len(events)}")
    print(f"  Dəyərli (premiumlu): {len(valued)}  ·  ümumi premium: {total} {ccy}")
    print(f"  Dataset           : {dataset} ({ds_kind})")
    if warnings:
        print(f"  Xəbərdarlıq ({len(warnings)}):")
        for w in warnings[:12]:
            print(f"     · {w}")
        if len(warnings) > 12:
            print(f"     … +{len(warnings) - 12} daha")
    if not events:
        print("\n  Göndəriləcək hadisə yoxdur.\n")
        return 0

    sample = events[0]
    cd = sample["custom_data"]
    val = f"{cd['value']} {cd['currency']}" if "value" in cd else "(dəyərsiz)"
    print("\n  Nümunə hadisə (ilk sətir, PII hash-lənib):")
    print(f"     event={sample['event_name']} id={sample['event_id']} value={val}")
    print(f"     user_data açarları: {sorted(sample['user_data'])}")

    test_code = args.test_code if args.test_code is not None else config.TEST_EVENT_CODE

    if not args.send:
        print("\n  DRY-RUN — heç nə göndərilmədi. Göndərmək üçün --send əlavə et.")
        print("  Əvvəlcə Test Events ilə yoxlamaq tövsiyə olunur:")
        print("     import_sales.py <fayl> --send --test-code TESTxxxxx\n")
        return 0

    sent = 0
    try:
        for batch in _chunks(events, 1000):  # Meta limit: 1000 events / request
            resp = capi.send_events(batch, dataset_id=dataset, test_event_code=test_code)
            sent += int(resp.get("events_received", 0))
            print(f"  → göndərildi: events_received={resp.get('events_received')} "
                  f"fbtrace_id={resp.get('fbtrace_id')}")
    except Exception as exc:
        print(f"\n  ✗ Göndərmə dayandı ({sent} göndərildikdən sonra): {exc}\n")
        return 3

    print("\n" + "=" * 64)
    where = "Test Events" if test_code else "Overview (production)"
    print(f"  ✓ {sent} hadisə qəbul edildi — Events Manager → {where}-də görünəcək.")
    print("=" * 64 + "\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
